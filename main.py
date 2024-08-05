"""
sample1:从领星下载广告-广告asin分析，选择ppt模版
sample2:领星-统计-订单利润，选择ppt模版
sample3:统计-产品表现，选择ppt模版
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string

# make the worksheet for totality
wb = Workbook()
sheet_total = wb.active
sheet_total.title = 'total'

# insert column titles
column_titles = ['ASIN', '品名', '价格', '曝光', '点击', '广告费',
                 '店铺销售额', '广告销售额', '店铺销量', '广告销量',
                 '退货量', '退款量',
                 '毛利润', '毛利率', '排名']

# copy titles to the sheet_total
def copy_titles(titles_list, sheet):
    """
    copy titles from the titles list to the sheet.

    :param titles_list: list. a list of titles
    :param sheet:variable. the sheet name of the target sheet
    :return: 'Done' when all titles have been copied.
    """
    column_titles_len = len(titles_list)
    count = 0
    for row in sheet.iter_rows(min_row=1,
                               max_col=column_titles_len):
        for cell in row:
            if count < column_titles_len:
                cell.value = column_titles[count]
            count += 1
    return 'Done'

# fill sheet totals
copy_titles(column_titles, sheet_total)

# clean up source 1 sheet
source1 = load_workbook('sample1.xlsx')
source_sheet1 = source1.active
source_sheet1.delete_rows(2, amount=1)


# make a function to copy from source1 to sheet_total
def copy_source1(source_row_num, dest_col):
    """
    copy from the source1 sheet to the target sheet.
    :param source_row_num: int. row number of the source sheet to
    copy from.
    :param dest_col: int. column number of the target sheet.
    :return: None
    """
    list_source = []
    for i, row in enumerate(source_sheet1):
        if i == 0:
            continue
        item = row[source_row_num].value
        list_source.append(item)
    list_len = len(list_source)
    i = 0
    for row in sheet_total[f"{dest_col}2:{dest_col}{list_len + 1}"]:
        for cell in row:
            cell.value = list_source[i]
            i += 1

# a function to copy source1 to sheet total
# copy asin source_col=3,
# copy 价格/竞价 col=7
# copy 曝光 col=9
# copy 点击 col=10
# copy 广告花费 col=13
# copy 店铺销售额 col=14
# copy 广告销售额 col=15
# copy 店铺销量 col=23
# copy 广告销量 col=24
def fill_fr_source1():
    """
    fill the target sheet with data from source sheet1.
    """
    source1_dict = {3: 'A', 7: 'C', 9: 'D', 10: 'E', 13: 'F', 14: 'G',
                    15: 'H', 23: 'I',
                    24: 'J'}
    for key, value in source1_dict.items():
        copy_source1(key, value)


fill_fr_source1()

#load source 2 sheet
source2=load_workbook('sample2.xlsx')
s = source2.active
source_sheet2=source2.active

def make_dict_source(source_sheet, lookup_col_num, return_col_num):
    """
    create a dictionary with the data from the source sheet's column

    :param source_sheet:variable of a source sheet.
    :param lookup_col_num:string. keys of the dictionary
    :param return_col_num:string. values of the dictionary
    :return:a dictionary with the keys of the lookup column and
    values of the return column.
    """
    dict = {}
    for i, row in enumerate(source_sheet):
        if i == 0:
            continue
        key = row[lookup_col_num].value
        value = row[return_col_num].value
        dict[key] = value

    return dict


# print(make_dict_source2(1,3))

# fill the sheet_total with the dict from source2
def fill_sheet_total(source_dict,
                     lookup_col_num,
                     target_col_num):
    """
    fill the target sheet with data from the dictionary made from
    source sheet if the lookup column number value is None. Get the
    value from the dictionary and add it to the cell.
    :param source_dict: a dictionary
    :param lookup_col_num: int. the target column number of the lookup column
    :param target_col_num: int. the target column number of the
    cell to be filled in.
    :return: None
    """
    source_dict = source_dict
    for row in sheet_total:
        if row[target_col_num].value == None:
            row[target_col_num].value = source_dict.get(
                row[lookup_col_num].value)


# fill 品名
fill_sheet_total(make_dict_source(source_sheet2, column_index_from_string('A')-1,
                                  column_index_from_string('B')-1),
                 0, column_index_from_string('B')-1)

# fill 退货量
fill_sheet_total(make_dict_source(source_sheet2,
                                  column_index_from_string('A')-1,
                                  column_index_from_string('J')-1),
                 0, column_index_from_string('K')-1)

# fill 退款量
fill_sheet_total(make_dict_source(source_sheet2, column_index_from_string('A')-1,
                                  column_index_from_string('K')-1),
                 0, column_index_from_string('L')-1)

# fill 利润
fill_sheet_total(make_dict_source(source_sheet2, column_index_from_string('A')-1,
                                  column_index_from_string('D')-1),
                 0, column_index_from_string('M')-1)

# fill 毛利率
fill_sheet_total(make_dict_source(source_sheet2, column_index_from_string('A')-1,
                                  column_index_from_string('E')-1),
                 0, column_index_from_string('N')-1)

# load source sheet 3
source3 = load_workbook('sample3.xlsx')
source_sheet3 = source3.active


def edit_source_dict(source_dict):
    """
    edit the values of the dictionary to save only the number part.
    :param source_dict: dict. the dictionary to be edited.
    :return: the edited dictionary
    """
    ready_dict = source_dict
    modified_dict = {}
    for key, value in ready_dict.items():
        if value == None:
            modified_dict[key] = 0
        else:
            modified_dict[key] = value.split('：')[-1]

    return modified_dict


# fill 排名
fill_sheet_total(edit_source_dict(make_dict_source(source_sheet3,
                                                   column_index_from_string('B')-1,
                                  column_index_from_string(
                                      'C')-1)), 0,
                 column_index_from_string('O')-1)


# total sheet for myyweld
sheet_mw = wb.create_sheet("MW")
# total sheet for ttamplar
sheet_tp = wb.create_sheet("TP")

# copy titles to two sheets
copy_titles(column_titles, sheet_mw)
copy_titles(column_titles, sheet_tp)


# sort the data to the two sheets function
def sort_data():
    """
    if the row 1 cell contains 'W' then copy it to the myyweld
    sheet else copy it to the ttamplar sheet.
    :return: None
    """
    for i, row in enumerate(sheet_total):
        if i == 0:
            continue
        if 'W' in row[1].value:
            mw_max_row = sheet_mw.max_row + 1
            for cell in row:
                sheet_mw.cell(row=mw_max_row,
                              column=cell.column).value = cell.value
        else:
            tp_max_row = sheet_tp.max_row + 1
            for cell in row:
                sheet_tp.cell(row=tp_max_row,
                              column=cell.column).value = cell.value


# sort and fill the two brand sheets.
sort_data()


# sum up the two brand sheet.
def sum_up(sheet):
    """
    sum up some columns
    :param sheet: worksheet to work on.
    :return: None
    """
    max_row = sheet.max_row
    sheet.cell(row=max_row + 1, column=2).value = '合计'
    for cell in sheet[max_row + 1]:
        if 3 < cell.column < 15:
            column_letter = get_column_letter(cell.column)
            cell.value = (f"=SUM({column_letter}2:{column_letter}"
                          f"{max_row})")


# sum up for two brand sheets
sum_up(sheet_mw)
sum_up(sheet_tp)



# save the final sheet_total
wb.save('total.xlsx')